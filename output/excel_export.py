"""
Excel 计算书导出模块

使用 openpyxl 生成格式化的 .xlsx 计算书。
"""

from io import BytesIO
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import (Font, Alignment, Border, Side, PatternFill,
                              numbers)
from openpyxl.utils import get_column_letter

from core.current_calc import CurrentResult
from core.cable_select import CableSelectionResult
from core.voltage_drop import VoltageDropResult
from core.short_circuit import ShortCircuitResult


# 样式定义
TITLE_FONT = Font(name="微软雅黑", size=16, bold=True, color="FFFFFF")
HEADER_FONT = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
NORMAL_FONT = Font(name="微软雅黑", size=11, color="333333")
RESULT_FONT = Font(name="微软雅黑", size=12, bold=True, color="C00000")
PASS_FONT = Font(name="微软雅黑", size=11, bold=True, color="007A33")
FAIL_FONT = Font(name="微软雅黑", size=11, bold=True, color="FF0000")

TITLE_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FILL = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
ALT_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
PASS_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
FAIL_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)

CENTER_ALIGN = Alignment(horizontal="center", vertical="center")
LEFT_ALIGN = Alignment(horizontal="left", vertical="center")


def _style_header_row(ws, row: int, max_col: int):
    """设置表头行样式"""
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER


def _style_data_row(ws, row: int, max_col: int, alt: bool = False):
    """设置数据行样式"""
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = NORMAL_FONT
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER
        if alt:
            cell.fill = ALT_FILL


def _add_title_row(ws, row: int, title: str, max_col: int):
    """添加标题行"""
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max_col)
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = TITLE_FONT
    cell.fill = TITLE_FILL
    cell.alignment = CENTER_ALIGN
    return row + 1


def _add_section_title(ws, row: int, title: str, max_col: int):
    """添加节标题"""
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max_col)
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = Font(name="微软雅黑", size=13, bold=True, color="1F4E79")
    cell.fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    cell.alignment = LEFT_ALIGN
    return row + 1


def _write_kv_rows(ws, start_row: int, data: list, max_col: int = 3):
    """写入键值对行"""
    row = start_row
    for i, (label, value) in enumerate(data):
        ws.cell(row=row, column=1, value=label)
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=max_col)
        ws.cell(row=row, column=2, value=value)
        _style_data_row(ws, row, max_col, alt=(i % 2 == 0))
        row += 1
    return row


def export_excel(current_result: CurrentResult,
                 cable_result: CableSelectionResult,
                 vd_result: VoltageDropResult,
                 sc_result: ShortCircuitResult,
                 filepath: str = None) -> str:
    """
    导出完整计算书到 Excel。

    参数:
        current_result: 电流计算结果
        cable_result:  电缆选型结果
        vd_result:      压降计算结果
        sc_result:      短路校验结果
        filepath:       保存路径，为 None 时返回默认路径

    返回:
        保存的文件路径
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "电气计算书"
    max_col = 5

    # 列宽设置
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 18

    row = 1

    # 封面标题
    row = _add_title_row(ws, row,
                         f"电气设计计算书  —  {datetime.now().strftime('%Y-%m-%d')}",
                         max_col)

    # 项目信息
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max_col)
    row += 1
    info_data = [
        ("设计依据", "GB 50217-2018 / GB 50054-2011"),
        ("计算日期", datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("软件版本", "电气计算工具 v1.0"),
    ]
    row = _write_kv_rows(ws, row + 1, info_data, max_col)
    row += 1

    # ========================================
    # ① 电流计算
    # ========================================
    row = _add_section_title(ws, row, "① 电流计算", max_col)
    row += 1

    headers = ["参数", "符号", "数值", "单位", "备注"]
    for ci, h in enumerate(headers, 1):
        ws.cell(row=row, column=ci, value=h)
    _style_header_row(ws, row, len(headers))
    row += 1

    phase_str = "三相" if current_result.is_three_phase else "单相"
    current_data = [
        ("相别", "—", phase_str, "—", ""),
        ("线电压", "U_n", f"{current_result.voltage:.0f}", "V", ""),
        ("功率因数", "cosφ", f"{current_result.cos_phi:.2f}", "—", ""),
        ("有功功率", "P_js", f"{current_result.p_js:.2f}", "kW", "P_e × Kx"),
        ("无功功率", "Q_js", f"{current_result.q_js:.2f}", "kvar", "P_js × tanφ"),
        ("视在功率", "S_js", f"{current_result.s_js:.2f}", "kVA", "P_js / cosφ"),
        ("计算电流", "I_js", f"{current_result.i_js:.2f}", "A",
         "P_js×1000/(√3×U_n×cosφ)"),
    ]

    for i, (param, symbol, value, unit, note) in enumerate(current_data):
        ws.cell(row=row, column=1, value=param)
        ws.cell(row=row, column=2, value=symbol)
        ws.cell(row=row, column=3, value=value)
        ws.cell(row=row, column=4, value=unit)
        ws.cell(row=row, column=5, value=note)
        font = RESULT_FONT if param == "计算电流" else NORMAL_FONT
        for c in range(1, 6):
            ws.cell(row=row, column=c).font = font
        _style_data_row(ws, row, 5, alt=(i % 2 == 0))
        row += 1

    row += 1

    # ========================================
    # ② 电缆选型
    # ========================================
    row = _add_section_title(ws, row, "② 电缆选型 (GB 50217)", max_col)
    row += 1

    inp = cable_result.input
    k = cable_result.correction_factors

    # 输入参数
    input_headers = ["项目", "内容", "备注"]
    for ci, h in enumerate(input_headers, 1):
        ws.cell(row=row, column=ci, value=h)
    _style_header_row(ws, row, 3)
    row += 1

    input_data = [
        ("导体材料", inp.conductor_cn, ""),
        ("绝缘类型", inp.insulation_cn, ""),
        ("敷设方式", inp.method_cn, ""),
        ("电压等级", inp.voltage_level, ""),
        ("环境温度", f"{inp.ambient_temp:.0f} °C", ""),
        ("并列回路数", f"{inp.num_circuits} 回", ""),
        ("计算电流 I_js", f"{inp.i_js:.2f} A", ""),
    ]
    row = _write_kv_rows(ws, row, input_data, 3)
    row += 1

    # 修正系数
    factor_headers = ["修正系数", "符号", "数值"]
    for ci, h in enumerate(factor_headers, 1):
        ws.cell(row=row, column=ci, value=h)
    _style_header_row(ws, row, 3)
    row += 1

    factor_data = [
        ("温度修正", "Kθ", f"{k.get('k_temp', 0):.3f}"),
        ("敷设修正", "K₁", f"{k.get('k_method', 0):.3f}"),
        ("并列修正", "K₂", f"{k.get('k_group', 0):.3f}"),
        ("综合修正", "K", f"{k.get('total', 0):.3f}"),
    ]
    row = _write_kv_rows(ws, row, factor_data, 3)
    row += 1

    # 选型结果
    result_headers = ["项目", "内容", "单位"]
    for ci, h in enumerate(result_headers, 1):
        ws.cell(row=row, column=ci, value=h)
    _style_header_row(ws, row, 3)
    row += 1

    if cable_result.is_valid:
        result_data = [
            ("推荐截面", cable_result.selected_section, "mm²"),
            ("基准载流量", f"{cable_result.raw_ampacity:.0f}", "A"),
            ("修正后载流量", f"{cable_result.corrected_ampacity:.0f}", "A"),
            ("判定", "✓ 合格", ""),
        ]
    else:
        result_data = [
            ("选型结果", "失败", ""),
            ("原因", cable_result.fail_reason, ""),
        ]

    row = _write_kv_rows(ws, row, result_data, 3)
    row += 1

    # ========================================
    # ③ 压降计算
    # ========================================
    row = _add_section_title(ws, row, "③ 压降校验 (GB 50054 §6.4)", max_col)
    row += 1

    vd_inp = vd_result.input
    vd_data = [
        ("回路类型", vd_inp.circuit_type, ""),
        ("电缆长度", f"{vd_inp.length_m:.0f}", "m"),
        ("电缆截面", f"{vd_inp.section}", "mm²"),
        ("线电压", f"{vd_inp.voltage:.0f}", "V"),
        ("功率因数", f"{vd_inp.cos_phi:.2f}", ""),
        ("单位电阻 R₀", f"{vd_result.r0:.4f}", "Ω/km"),
        ("单位电抗 X₀", f"{vd_result.x0:.4f}", "Ω/km"),
        ("压降 ΔU", f"{vd_result.delta_u_v:.2f}", "V"),
        ("压降百分比 ΔU%", f"{vd_result.delta_u_percent:.2f}%", ""),
        ("允许限值", f"{vd_result.limit:.1f}%", ""),
    ]
    row = _write_kv_rows(ws, row, vd_data, 3)

    # 判定
    vd_status = "✓ 合格" if vd_result.is_pass else "✗ 不合格"
    ws.cell(row=row, column=1, value="判定")
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
    cell = ws.cell(row=row, column=2, value=vd_status)
    cell.font = PASS_FONT if vd_result.is_pass else FAIL_FONT
    _style_data_row(ws, row, 3)
    row += 2

    # ========================================
    # ④ 短路热稳定校验
    # ========================================
    row = _add_section_title(ws, row, "④ 短路热稳定校验 (GB 50054 §6.2.3)", max_col)
    row += 1

    sc_inp = sc_result.input
    sc_data = [
        ("短路电流 I\"ₖ₃", f"{sc_inp.ik_a:.0f} A ({sc_inp.ik_a/1000:.2f} kA)", ""),
        ("持续时间 t", f"{sc_inp.t_s:.2f}", "s"),
        ("热稳定系数 K", f"{sc_result.k:.0f}", ""),
        ("要求最小截面 S_min", f"{sc_result.s_min:.1f}", "mm²"),
        ("已选截面", f"{sc_inp.selected_section}", "mm²"),
    ]
    row = _write_kv_rows(ws, row, sc_data, 3)

    sc_status = "✓ 合格" if sc_result.is_pass else f"✗ 不合格 (需 ≥ {sc_result.required_section} mm²)"
    ws.cell(row=row, column=1, value="判定")
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
    cell = ws.cell(row=row, column=2, value=sc_status)
    cell.font = PASS_FONT if sc_result.is_pass else FAIL_FONT
    _style_data_row(ws, row, 3)
    row += 2

    # ========================================
    # 最终结论
    # ========================================
    all_pass = (cable_result.is_valid and vd_result.is_pass and sc_result.is_pass)
    row = _add_section_title(ws, row, "最终结论", max_col)
    row += 1

    conclusion = "✅ 全部校验通过，电缆选型满足 GB 50217 / GB 50054 要求。" if all_pass \
        else "❌ 存在不合格项，请调整参数后重新计算。"

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max_col)
    cell = ws.cell(row=row, column=1, value=conclusion)
    cell.font = Font(name="微软雅黑", size=13, bold=True,
                     color="007A33" if all_pass else "FF0000")
    cell.alignment = CENTER_ALIGN
    fill = PASS_FILL if all_pass else FAIL_FILL
    for c in range(1, max_col + 1):
        ws.cell(row=row, column=c).fill = fill

    # 保存
    if filepath is None:
        filepath = f"电气计算书_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    wb.save(filepath)
    return filepath


def export_excel_simple(data_dict: dict, filepath: str = None) -> str:
    """简化版 Excel 导出，适用于只有部分计算结果的场景。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "电气计算书"

    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 28

    row = 1
    row = _add_title_row(ws, row, f"电气设计计算书  —  {datetime.now().strftime('%Y-%m-%d')}", 2)
    row += 1

    for section_title, items in data_dict.items():
        row = _add_section_title(ws, row, section_title, 2)
        row += 1

        for i, (label, value) in enumerate(items):
            ws.cell(row=row, column=1, value=label)
            ws.cell(row=row, column=2, value=str(value))
            _style_data_row(ws, row, 2, alt=(i % 2 == 0))
            row += 1
        row += 1

    if filepath is None:
        filepath = f"电气计算书_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    wb.save(filepath)
    return filepath
